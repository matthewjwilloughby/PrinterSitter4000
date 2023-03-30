from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

# from selenium.webdriver.common import keys
from selenium.webdriver.common.by import By
from openpyxl import Workbook, load_workbook
import time


# Install Webdriver
PATH = r"C:\Users\User\Desktop\Python\Chromedriver102\chromedriver.exe"
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))

# Load excel workbook 
wb = load_workbook(r'C:\Users\User\Desktop\Python\Printer-Sitter-4000\Printer Toner Levels.xlsx')
ws = wb.active

# Open printer URL and login
def login(printerName, url, username_radial_button_ID, username_element_ID, username, password_form_ID, password):
    try:
        # Load printers URL. 
        driver.get(url)

        # Login to printer WebUi if the printer needs a login
        # Username
        if username_element_ID != 'none':
            driver.find_element(by=By.NAME, value=username_radial_button_ID).click()
            login_username = driver.find_element(by=By.NAME, value=username_element_ID)
            login_username.send_keys(username)
        # Password 
        if password_form_ID != 'none':
            login = driver.find_element(by=By.NAME, value=password_form_ID)
            login.send_keys(password)
            login.send_keys(Keys.ENTER)
    except:
        print("Failed to load", printerName)

# Set excel column for the desired toner color. 
def excel_color_column(color):
    match color:
        case 'black':
            return(9)
        case 'cyan':
            return(10)
        case 'magenta':
            return(11)
        case 'yellow':
            return(12)
        case 'mkb1':
            return(13)
        case 'mkb2':
            return(14)

#Retrieve toner color level and input value into designated spreadsheet cell
def retrieve_toner_levels(printerName, excel_sheet_toner_cartrige_column, cartridge_XPath_dictionary):
        for attempt in range(1,5):
            try:
                for color in cartridge_XPath_dictionary:
                    column = excel_color_column(color)
                    toner_value = driver.find_element(by=By.XPATH, value=cartridge_XPath_dictionary[color])
                    try:
                        ws.cell(column=column, row=excel_sheet_toner_cartrige_column).value = int(toner_value.text[:-1])
                    except:
                        ws.cell(column=column, row=excel_sheet_toner_cartrige_column).value = str(toner_value.text)
                break
            except:
                print("Unable to retrieve toner levels for", printerName + ": Attempt", attempt)
                # Give the printer UI more time to load
                time.sleep(1) 


# Printer 01 - Samsung SCX-6545
printerName = 'Printer 01 - Samsung SCX-6545'
login( printerName ,'http://10.0.10.94/', 'none', 'none', 'none','none','none' )
retrieve_toner_levels(
    printerName,
    4,
    {
    'black':'/html/body/div[1]/div/div/div[2]/div/div/div/div/div/div[2]/div/div/div/div/div/div/div[2]/div[1]/div/div/div/div/div/form/fieldset[1]/div/div/div[1]/div/div/div/div[1]/div/div/div/div[2]/div[1]/div/div/div/div/div[1]/div[2]/div/div/table/tbody/tr/td[2]/div/div/div[2]'  
    }
)

# Pinter 02 - Canon MF731C
printerName = 'Printer 2 - Canon MF731C'
login(printerName, 'http://10.0.10.12/', 'none', 'none', 'none','i0019','Password' )
retrieve_toner_levels(
    printerName,
    5,
    {
        'black':'/html/body/div[1]/div[2]/div[2]/div/div[1]/div/div[2]/div[2]/div[2]/div[3]/table/tbody/tr[4]/td[2]',
        'cyan':'/html/body/div[1]/div[2]/div[2]/div/div[1]/div/div[2]/div[2]/div[2]/div[3]/table/tbody/tr[1]/td[2]',
        'magenta':'/html/body/div[1]/div[2]/div[2]/div/div[1]/div/div[2]/div[2]/div[2]/div[3]/table/tbody/tr[2]/td[2]',
        'yellow':'/html/body/div[1]/div[2]/div[2]/div/div[1]/div/div[2]/div[2]/div[2]/div[3]/table/tbody/tr[3]/td[2]'
    }
    )

# Printer 03 - Brother MFC-7860DW
printerName = 'Printer 3 - Brother MFC-7860DW'
login(printerName, 'http://10.0.10.56/', 'none', 'none', 'none','none','none' )
retrieve_toner_levels(
    printerName,
    6,
    {
        'black':'/html/body/table/tbody/tr/td/form/table[1]/tbody/tr[2]/td/table/tbody/tr[1]/td[1]/table/tbody/tr/td/table/tbody/tr[2]/td/table/tbody/tr[1]/td/a/tt'
    }
    )

# Printer 04 - Cannon MF644Cdw Megans's
printerName = "CRPrinter04 - Cannon MF644Cdw Megan' printer"
login(printerName,'http://10.0.10.272/', 'none', 'none', 'none','i0019','Password' )
retrieve_toner_levels(
    printerName,
    7,
    {
        'black':'/html/body/div[1]/div[2]/div[2]/div/div[1]/div/div[2]/div[2]/div[2]/div[3]/table/tbody/tr[4]/td[2]',
        'cyan':'/html/body/div[1]/div[2]/div[2]/div/div[1]/div/div[2]/div[2]/div[2]/div[3]/table/tbody/tr[1]/td[2]',
        'magenta':'/html/body/div[1]/div[2]/div[2]/div/div[1]/div/div[2]/div[2]/div[2]/div[3]/table/tbody/tr[2]/td[2]',
        'yellow':'/html/body/div[1]/div[2]/div[2]/div/div[1]/div/div[2]/div[2]/div[2]/div[3]/table/tbody/tr[3]/td[2]'
    }
    )

# Printer 05 - Cannon MF440
printerName = 'Printer 5 - Cannon MF440'
login(printerName,'http://10.0.10.141/', 'i0012', 'i0014', 'Username','i0016','Password' )
retrieve_toner_levels(
    printerName,
    8,
    {
        'black':'/html/body/div[1]/div[2]/div[2]/div/div[1]/div/div[2]/div[2]/div[2]/div[3]/table/tbody/tr/td[2]'
    }
    )

# Printer 06 - Cannon MF743Cdw
printerName = "CRPrinter06 - Cannon MF743Cdw"
login(printerName,'http://10.0.10.38/', 'none', 'none', 'none','i2101','Password' )
retrieve_toner_levels(
    printerName,
    9,
    {
        'black':'/html/body/div[1]/div[2]/div[2]/div/div[1]/div/div[2]/div[2]/div[2]/div[3]/table/tbody/tr[4]/td[2]',
        'cyan':'//*[@id="tonerInfomationModule"]/table/tbody/tr[1]/td[2]',
        'magenta':'/html/body/div[1]/div[2]/div[2]/div/div[1]/div/div[2]/div[2]/div[2]/div[3]/table/tbody/tr[2]/td[2]',
        'yellow':'/html/body/div[1]/div[2]/div[2]/div/div[1]/div/div[2]/div[2]/div[2]/div[3]/table/tbody/tr[3]/td[2]'
    }
    )

# Printer 07 - Brother MFC-J985DW

printerName = "Printer 07 - Brother MFC-J985DW"
login(printerName,'http://10.0.10.87/', 'none', 'none', 'none','i2101','Password' )
retrieve_toner_levels(
    printerName,
    10,
    {
        'black':'/html/body/div[1]/div[2]/div[2]/div/div[1]/div/div[2]/div[2]/div[2]/div[3]/table/tbody/tr[4]/td[2]',
        'cyan':'//*[@id="tonerInfomationModule"]/table/tbody/tr[1]/td[2]',
        'magenta':'/html/body/div[1]/div[2]/div[2]/div/div[1]/div/div[2]/div[2]/div[2]/div[3]/table/tbody/tr[2]/td[2]',
        'yellow':'/html/body/div[1]/div[2]/div[2]/div/div[1]/div/div[2]/div[2]/div[2]/div[3]/table/tbody/tr[3]/td[2]'
    }
    )

# Printer 08 - Plotter Cannon iPF605 
printerName = "Plotter - Cannon iPF605"
login(printerName,'http://10.0.10.42/', 'none', 'none', 'none','none','none' )
retrieve_toner_levels(
    printerName,
    11,
    {
        'black':'/html/body/form/div[4]/table[2]/tbody/tr[3]/td/table/tbody/tr[3]/td/table[2]/tbody/tr[1]/td[2]/table/tbody/tr[2]/td/table/tbody/tr[4]/td[3]',
        'cyan':'/html/body/form/div[4]/table[2]/tbody/tr[3]/td/table/tbody/tr[3]/td/table[2]/tbody/tr[1]/td[1]/table/tbody/tr[2]/td/table/tbody/tr[4]/td[1]',
        'magenta':'/html/body/form/div[4]/table[2]/tbody/tr[3]/td/table/tbody/tr[3]/td/table[2]/tbody/tr[1]/td[1]/table/tbody/tr[2]/td/table/tbody/tr[4]/td[2]',
        'yellow':'/html/body/form/div[4]/table[2]/tbody/tr[3]/td/table/tbody/tr[3]/td/table[2]/tbody/tr[1]/td[1]/table/tbody/tr[2]/td/table/tbody/tr[4]/td[3]',
        'mkb1':'/html/body/form/div[4]/table[2]/tbody/tr[3]/td/table/tbody/tr[3]/td/table[2]/tbody/tr[1]/td[2]/table/tbody/tr[2]/td/table/tbody/tr[4]/td[1]',
        'mkb2':'/html/body/form/div[4]/table[2]/tbody/tr[3]/td/table/tbody/tr[3]/td/table[2]/tbody/tr[1]/td[2]/table/tbody/tr[2]/td/table/tbody/tr[4]/td[2]'
    }
    )

# Printer 09
printerName = "Printer 09 - Cannon MF445dw"
login(printerName,'http://10.0.10.227/', 'i0012', 'i0014', 'Username','i0016','Password' )
retrieve_toner_levels(
    printerName,
    12,
    {
        'black':'/html/body/div[1]/div[2]/div[2]/div/div[1]/div/div[2]/div[2]/div[2]/div[3]/table/tbody/tr/td[2]'
    }
    )

# Printer 10
printerName = "Printer 10 - Cannon MF414dw"
login(printerName,'http://10.0.10.79/', 'none', 'none', 'none','i0019','Password' )
retrieve_toner_levels(
    printerName,
    13,
    {
        'black':'/html/body/div[1]/div[2]/div[2]/div/div[1]/div/div[2]/div[2]/div[2]/div[3]/table/tbody/tr/td[2]'
    }
    )

# Save workbook and close script
wb.save(r'C:\Users\User\Desktop\Python\Printer-Sitter-4000\Printer Toner Levels2.xlsx')
time.sleep(1)
driver.quit()
wb.close